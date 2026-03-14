from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent.parent
MERCADO_PAGO_DEMO_PATH = ROOT / "mercado_pago_demo.xlsx"
FINANZAS_DEMO_PATH = ROOT / "FINANZAS_demo.xlsx"

HEADERS = [
    "Fecha",
    "Tipo",
    "Categoria",
    "Subcategoria",
    "Descripcion",
    "Monto",
    "Cuenta",
    "Canal",
    "Mes",
    "Compartido",
]

ARS_FORMAT = '#,##0.00 "ARS"'
MONTH_FORMAT = r"mmm\-yyyy"


def build_mercado_pago_rows() -> list[dict[str, object]]:
    return [
        {"date": "03-01-2026", "type": "Transferencia recibida ingreso principal", "ref": "200001460001", "amount": "380.000,00", "balance": "505.000,00"},
        {"date": "04-01-2026", "type": "Pago mercado semanal", "ref": "200001460002", "amount": "-24.850,00", "balance": "480.150,00"},
        {"date": "05-01-2026", "type": "Pago combustible", "ref": "200001460003", "amount": "-19.200,00", "balance": "460.950,00"},
        {"date": "06-01-2026", "type": "Suscripcion media demo", "ref": "200001460004", "amount": "-8.499,00", "balance": "452.451,00"},
        {"date": "07-01-2026", "type": "Pago movilidad urbana", "ref": "200001460005", "amount": "-5.240,00", "balance": "447.211,00"},
        {"date": "08-01-2026", "type": "Transferencia recibida gastos compartidos", "ref": "200001460006", "amount": "22.000,00", "balance": "469.211,00"},
        {"date": "09-01-2026", "type": "Pago salud y cuidado", "ref": "200001460007", "amount": "-11.780,00", "balance": "457.431,00"},
        {"date": "10-01-2026", "type": "Rendimientos", "ref": "200001460008", "amount": "198,00", "balance": "457.629,00"},
        {"date": "12-01-2026", "type": "Transferencia enviada vivienda mensual", "ref": "200001460009", "amount": "-145.000,00", "balance": "312.629,00"},
        {"date": "13-01-2026", "type": "Pago salida casual", "ref": "200001460010", "amount": "-13.450,00", "balance": "299.179,00"},
        {"date": "15-01-2026", "type": "Pago compra de cercania", "ref": "200001460011", "amount": "-17.920,00", "balance": "281.259,00"},
        {"date": "18-01-2026", "type": "Transferencia recibida proyecto independiente", "ref": "200001460012", "amount": "95.000,00", "balance": "376.259,00"},
        {"date": "20-01-2026", "type": "Pago actividad fisica", "ref": "200001460013", "amount": "-14.000,00", "balance": "362.259,00"},
        {"date": "22-01-2026", "type": "Pago mercado semanal", "ref": "200001460014", "amount": "-28.640,00", "balance": "333.619,00"},
        {"date": "25-01-2026", "type": "Transferencia recibida ajuste demo", "ref": "200001460015", "amount": "8.750,00", "balance": "342.369,00"},
        {"date": "28-01-2026", "type": "Transferencia enviada fondo viaje", "ref": "200001460016", "amount": "-35.000,00", "balance": "307.369,00"},
        {"date": "01-02-2026", "type": "Transferencia recibida ingreso principal", "ref": "200001460017", "amount": "390.000,00", "balance": "697.369,00"},
        {"date": "02-02-2026", "type": "Pago mercado semanal", "ref": "200001460018", "amount": "-26.480,00", "balance": "670.889,00"},
        {"date": "03-02-2026", "type": "Pago combustible", "ref": "200001460019", "amount": "-21.600,00", "balance": "649.289,00"},
        {"date": "05-02-2026", "type": "Pago movilidad urbana", "ref": "200001460020", "amount": "-6.100,00", "balance": "643.189,00"},
        {"date": "06-02-2026", "type": "Transferencia recibida gastos compartidos", "ref": "200001460021", "amount": "24.500,00", "balance": "667.689,00"},
        {"date": "07-02-2026", "type": "Pago salud y cuidado", "ref": "200001460022", "amount": "-9.980,00", "balance": "657.709,00"},
        {"date": "09-02-2026", "type": "Pago servicio conectividad", "ref": "200001460023", "amount": "-18.500,00", "balance": "639.209,00"},
        {"date": "10-02-2026", "type": "Rendimientos", "ref": "200001460024", "amount": "245,00", "balance": "639.454,00"},
        {"date": "12-02-2026", "type": "Transferencia enviada vivienda mensual", "ref": "200001460025", "amount": "-150.000,00", "balance": "489.454,00"},
        {"date": "14-02-2026", "type": "Pago salida casual", "ref": "200001460026", "amount": "-16.800,00", "balance": "472.654,00"},
        {"date": "17-02-2026", "type": "Pago compra de cercania", "ref": "200001460027", "amount": "-19.460,00", "balance": "453.194,00"},
        {"date": "20-02-2026", "type": "Transferencia recibida proyecto independiente", "ref": "200001460028", "amount": "110.000,00", "balance": "563.194,00"},
        {"date": "21-02-2026", "type": "Pago regalo planificado", "ref": "200001460029", "amount": "-42.000,00", "balance": "521.194,00"},
        {"date": "23-02-2026", "type": "Pago mercado semanal", "ref": "200001460030", "amount": "-31.540,00", "balance": "489.654,00"},
        {"date": "25-02-2026", "type": "Transferencia recibida ajuste viaje demo", "ref": "200001460031", "amount": "12.400,00", "balance": "502.054,00"},
        {"date": "27-02-2026", "type": "Pago actividad fisica", "ref": "200001460032", "amount": "-14.000,00", "balance": "488.054,00"},
    ]


def build_finanzas_history() -> list[tuple[date, str, str, str, str, float, str, str, str, str]]:
    return [
        (date(2025, 11, 2), "Ingreso", "Ingresos", "Base", "Ingreso principal demo", 355000.0, "Banco", "Transferencia", "", "No"),
        (date(2025, 11, 3), "Gasto", "Comida", "Mercado", "Compra semanal demo", 42350.0, "Banco", "Debito", "", "No"),
        (date(2025, 11, 5), "Gasto", "Vivienda", "Fijo", "Pago vivienda demo", 138000.0, "Banco", "Transferencia", "", "No"),
        (date(2025, 11, 8), "Gasto", "Transporte", "Movilidad", "Carga movilidad demo", 17400.0, "Mercado Pago", "QR", "", "No"),
        (date(2025, 11, 10), "Ingreso", "Ingresos", "Proyecto", "Cobro proyecto demo", 92000.0, "Mercado Pago", "Transferencia", "", "No"),
        (date(2025, 11, 14), "Gasto", "Comida", "Salida", "Comida compartida demo", 16800.0, "Mercado Pago", "QR", "", "Si"),
        (date(2025, 11, 18), "Gasto", "Salud", "Cuidado", "Compra salud demo", 12350.0, "Mercado Pago", "Tarjeta", "", "No"),
        (date(2025, 11, 27), "Gasto", "Servicios", "Conectividad", "Servicio hogar demo", 18500.0, "Banco", "Debito", "", "No"),
        (date(2025, 12, 2), "Ingreso", "Ingresos", "Base", "Ingreso principal demo", 360000.0, "Banco", "Transferencia", "", "No"),
        (date(2025, 12, 4), "Gasto", "Comida", "Mercado", "Compra semanal demo", 27840.0, "Mercado Pago", "QR", "", "No"),
        (date(2025, 12, 6), "Gasto", "Transporte", "Movilidad", "Carga movilidad demo", 19890.0, "Mercado Pago", "QR", "", "No"),
        (date(2025, 12, 9), "Ingreso", "Ingresos", "Ajuste", "Reintegro demo", 12450.0, "Mercado Pago", "Transferencia", "", "No"),
        (date(2025, 12, 11), "Gasto", "Vivienda", "Fijo", "Pago vivienda demo", 140000.0, "Banco", "Transferencia", "", "No"),
        (date(2025, 12, 14), "Gasto", "Ocio", "Digital", "Suscripcion demo", 8499.0, "Mercado Pago", "Debito", "", "No"),
        (date(2025, 12, 18), "Ingreso", "Ingresos", "Proyecto", "Cobro proyecto demo", 105000.0, "Mercado Pago", "Transferencia", "", "No"),
        (date(2025, 12, 20), "Gasto", "Comida", "Salida", "Cena demo", 24600.0, "Mercado Pago", "QR", "", "Si"),
        (date(2025, 12, 23), "Gasto", "Regalos", "Planificado", "Regalo demo", 56400.0, "Mercado Pago", "QR", "", "No"),
        (date(2026, 1, 2), "Ingreso", "Ingresos", "Base", "Ingreso principal demo", 375000.0, "Banco", "Transferencia", "", "No"),
        (date(2026, 1, 4), "Gasto", "Comida", "Mercado", "Compra semanal demo", 23120.0, "Mercado Pago", "QR", "", "No"),
        (date(2026, 1, 6), "Gasto", "Transporte", "Movilidad", "Carga movilidad demo", 18800.0, "Mercado Pago", "QR", "", "No"),
        (date(2026, 1, 8), "Ingreso", "Ingresos", "Compartido", "Reintegro compartido demo", 22000.0, "Mercado Pago", "Transferencia", "", "Si"),
        (date(2026, 1, 10), "Gasto", "Vivienda", "Fijo", "Pago vivienda demo", 145000.0, "Banco", "Transferencia", "", "No"),
        (date(2026, 1, 13), "Gasto", "Comida", "Salida", "Salida demo", 13450.0, "Mercado Pago", "QR", "", "No"),
        (date(2026, 1, 18), "Ingreso", "Ingresos", "Proyecto", "Cobro proyecto demo", 95000.0, "Mercado Pago", "Transferencia", "", "No"),
        (date(2026, 1, 20), "Gasto", "Bienestar", "Habito", "Actividad fisica demo", 14000.0, "Mercado Pago", "Debito", "", "No"),
        (date(2026, 1, 22), "Gasto", "Comida", "Mercado", "Compra semanal demo", 28640.0, "Mercado Pago", "QR", "", "No"),
        (date(2026, 1, 25), "Ingreso", "Ingresos", "Ajuste", "Ajuste demo", 8750.0, "Mercado Pago", "Transferencia", "", "No"),
    ]


def style_sheet(ws, title: str, subtitle: str) -> None:
    dark_fill = PatternFill("solid", fgColor="0F172A")
    accent_fill = PatternFill("solid", fgColor="0F766E")
    soft_fill = PatternFill("solid", fgColor="ECFDF5")
    header_fill = PatternFill("solid", fgColor="DFF7F3")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=14)
    thin = Side(style="thin", color="D9E2EC")

    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    ws["A1"] = title
    ws["A2"] = subtitle
    ws["A1"].fill = dark_fill
    ws["A2"].fill = accent_fill
    ws["A1"].font = title_font
    ws["A2"].font = white_font
    ws["A1"].alignment = Alignment(horizontal="left")
    ws["A2"].alignment = Alignment(horizontal="left")

    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=4, column=col).fill = header_fill
        ws.cell(row=4, column=col).font = Font(bold=True, color="0F172A")
        ws.cell(row=4, column=col).border = Border(top=thin, bottom=thin)

    widths = {
        "A": 13,
        "B": 12,
        "C": 18,
        "D": 18,
        "E": 30,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 12,
        "J": 11,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(1, ws.max_row + 1):
        if row >= 5:
            ws.cell(row=row, column=1).number_format = "dd/mm/yyyy"
            ws.cell(row=row, column=6).number_format = ARS_FORMAT
            ws.cell(row=row, column=9).number_format = MONTH_FORMAT
        if row in (1, 2):
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = dark_fill if row == 1 else accent_fill
                ws.cell(row=row, column=col).font = white_font
        if row == 3:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = soft_fill


def create_finanzas_demo() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Control de ingresos y gastos"

    style_sheet(
        ws,
        title="FINANZAS Demo | Base de trabajo para capturas",
        subtitle="Archivo sintetico y seguro para mostrar importaciones, dedupe y lectura mensual.",
    )

    for idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=4, column=idx, value=header)

    history = build_finanzas_history()
    for row_idx, row in enumerate(history, start=5):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        ws.cell(row=row_idx, column=9, value=f"=DATE(YEAR(A{row_idx}),MONTH(A{row_idx}),1)")
        ws.cell(row=row_idx, column=1).number_format = "dd/mm/yyyy"
        ws.cell(row=row_idx, column=6).number_format = ARS_FORMAT
        ws.cell(row=row_idx, column=9).number_format = MONTH_FORMAT

    last_row = 4 + len(history)
    table = Table(displayName="tblControlIngresosGastos", ref=f"A4:J{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    for row_idx in range(5, last_row + 1):
        ws.cell(row=row_idx, column=1).number_format = "dd/mm/yyyy"
        ws.cell(row=row_idx, column=6).number_format = ARS_FORMAT
        ws.cell(row=row_idx, column=9).number_format = MONTH_FORMAT

    summary = wb.create_sheet("Resumen")
    summary.sheet_view.showGridLines = False
    summary["A1"] = "Resumen Demo"
    summary["A2"] = "Workbook sintetico listo para importar desde la app."
    summary["A1"].font = Font(size=16, bold=True, color="0F172A")
    summary["A2"].font = Font(size=11, color="475569")
    cards = [
        ("Movimientos base", len(history), "Antes de importar Mercado Pago"),
        ("Periodo", "11/2025 a 01/2026", "Sirve para comparar enero contra febrero"),
        ("Hoja operativa", "Control de ingresos y gastos", "Con tabla tblControlIngresosGastos"),
        ("Estado", "Listo para demo", "Seguro para capturas y pruebas"),
    ]
    fills = ["DFF7F3", "E0F2FE", "FEF3C7", "F3E8FF"]
    for idx, (title, value, detail) in enumerate(cards, start=0):
        row = 4
        col = 1 + (idx * 3)
        summary.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        summary.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        summary.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 1)
        for r in range(row, row + 3):
            for c in range(col, col + 2):
                cell = summary.cell(r, c)
                cell.fill = PatternFill("solid", fgColor=fills[idx])
                cell.border = Border(
                    left=Side(style="thin", color="CBD5E1"),
                    right=Side(style="thin", color="CBD5E1"),
                    top=Side(style="thin", color="CBD5E1"),
                    bottom=Side(style="thin", color="CBD5E1"),
                )
                cell.alignment = Alignment(horizontal="center", vertical="center")
        summary.cell(row, col, title).font = Font(bold=True, color="0F172A")
        summary.cell(row + 1, col, value).font = Font(size=13, bold=True, color="0F172A")
        summary.cell(row + 2, col, detail).font = Font(size=10, color="475569")
    for col_letter in "ABCDEFGHJKLM":
        summary.column_dimensions[col_letter].width = 16

    wb.save(FINANZAS_DEMO_PATH)


def create_mercado_pago_demo() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet0"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "INITIAL_BALANCE"
    ws["B1"] = "CREDITS"
    ws["C1"] = "DEBITS"
    ws["D1"] = "FINAL_BALANCE"
    ws["A2"] = "125.000,00"
    ws["B2"] = "1.042.895,00"
    ws["C2"] = "-679.841,00"
    ws["D2"] = "488.054,00"

    headers = ["RELEASE_DATE", "TRANSACTION_TYPE", "REFERENCE_ID", "TRANSACTION_NET_AMOUNT", "PARTIAL_BALANCE"]
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=4, column=idx, value=header)

    rows = build_mercado_pago_rows()
    for row_idx, item in enumerate(rows, start=5):
        ws.cell(row=row_idx, column=1, value=item["date"])
        ws.cell(row=row_idx, column=2, value=item["type"])
        ws.cell(row=row_idx, column=3, value=item["ref"])
        ws.cell(row=row_idx, column=4, value=item["amount"])
        ws.cell(row=row_idx, column=5, value=item["balance"])

    fills = {
        1: PatternFill("solid", fgColor="0F172A"),
        2: PatternFill("solid", fgColor="DFF7F3"),
        4: PatternFill("solid", fgColor="CCFBF1"),
    }
    for row in (1, 2, 4):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if row in fills:
                cell.fill = fills[row]
            cell.font = Font(bold=True, color="FFFFFF" if row == 1 else "0F172A")
            cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18

    wb.save(MERCADO_PAGO_DEMO_PATH)


def main() -> None:
    create_mercado_pago_demo()
    create_finanzas_demo()
    print(f"Generated: {MERCADO_PAGO_DEMO_PATH.name}")
    print(f"Generated: {FINANZAS_DEMO_PATH.name}")


if __name__ == "__main__":
    main()
