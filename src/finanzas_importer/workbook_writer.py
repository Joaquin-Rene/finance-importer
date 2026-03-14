from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import unicodedata
from typing import NamedTuple
from tempfile import gettempdir

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from copy import copy

from .utils import create_backup_path, parse_number_ar

TABLE_NAME = "tblControlIngresosGastos"
LEGACY_TABLE_NAMES = ("tblJoaquin",)
SHEET_NAME = "Control de ingresos y gastos"
LEGACY_SHEET_NAMES = ("Joaquin",)
DEFAULT_ARS_FORMAT = '#,##0.00 "ARS"'
DEFAULT_DATE_FORMAT = "dd/mm/yyyy"
DEFAULT_MES_FORMAT = r"mmm\-yyyy"

DATE_FILTER_MODE_AFTER_MAX = "after_max_date"
DATE_FILTER_MODE_SKIP_EXISTING = "skip_existing_dates"


@dataclass
class ImportResult:
    added_rows: int
    duplicate_rows_mp_ref: int
    duplicate_rows_compound_key: int
    skipped_rows_by_date: int
    backup_path: Path | None
    note: str = ""


@dataclass
class ImportPlan:
    to_import_df: pd.DataFrame
    filtered_by_date_df: pd.DataFrame
    duplicate_mp_ref_df: pd.DataFrame
    duplicate_compound_df: pd.DataFrame
    existing_refs: set[str]
    total_input_rows: int
    table_name: str
    date_filter_mode: str


class TargetRange(NamedTuple):
    start_col: int
    start_row: int
    end_col: int
    end_row: int
    table_name: str
    has_table: bool


def resolve_sheet_name(workbook) -> str | None:
    for candidate in (SHEET_NAME, *LEGACY_SHEET_NAMES):
        if candidate in workbook.sheetnames:
            return candidate
    return None


def _get_table(sheet) -> Table | None:
    for table in sheet.tables.values():
        if table.name in (TABLE_NAME, *LEGACY_TABLE_NAMES):
            return table
    first_table = next(iter(sheet.tables.values()), None)
    return first_table


def _col_to_index(col_letters: str) -> int:
    result = 0
    for ch in col_letters:
        result = result * 26 + (ord(ch.upper()) - ord("A") + 1)
    return result


def _index_to_col(col_index: int) -> str:
    letters = ""
    while col_index > 0:
        col_index, remainder = divmod(col_index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _parse_table_ref(table_ref: str) -> tuple[int, int, int, int]:
    # Ejemplo: A1:K20
    start, end = table_ref.split(":")
    start_col = "".join(ch for ch in start if ch.isalpha())
    start_row = int("".join(ch for ch in start if ch.isdigit()))
    end_col = "".join(ch for ch in end if ch.isalpha())
    end_row = int("".join(ch for ch in end if ch.isdigit()))
    return _col_to_index(start_col), start_row, _col_to_index(end_col), end_row


def _resolve_target_range(ws) -> TargetRange:
    table = _get_table(ws)
    if table is not None:
        start_col, start_row, end_col, end_row = _parse_table_ref(table.ref)
        return TargetRange(
            start_col=start_col,
            start_row=start_row,
            end_col=end_col,
            end_row=end_row,
            table_name=table.name,
            has_table=True,
        )

    # Fallback sin tabla: asume estructura tabular desde A con encabezado en fila 1.
    start_col = 1
    start_row = 1
    end_col = start_col
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=start_row, column=col).value not in (None, ""):
            end_col = col
    last_data_row = start_row
    for row in range(ws.max_row, start_row, -1):
        if any(ws.cell(row=row, column=col).value not in (None, "") for col in range(start_col, end_col + 1)):
            last_data_row = row
            break

    return TargetRange(
        start_col=start_col,
        start_row=start_row,
        end_col=end_col,
        end_row=last_data_row,
        table_name=f"(sin tabla, rango A:{_index_to_col(end_col)})",
        has_table=False,
    )


def _normalize_header_name(value: object) -> str:
    return _strip_accents(str(value).strip().lower())


def _build_header_map(sheet, start_row: int, start_col: int, end_col: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(start_col, end_col + 1):
        header = sheet.cell(row=start_row, column=col).value
        if header in (None, ""):
            continue
        headers[_normalize_header_name(header)] = col
    return headers


def _extract_existing_refs(sheet, start_row: int, end_row: int, notes_col: int | None) -> set[str]:
    if notes_col is None:
        return set()

    refs: set[str] = set()
    pattern = re.compile(r"mp_ref=([^;\s]+)")

    for row in range(start_row + 1, end_row + 1):
        note = sheet.cell(row=row, column=notes_col).value
        if not note:
            continue
        m = pattern.search(str(note))
        if m:
            refs.add(m.group(1).strip())
    return refs


def _extract_existing_dates(sheet, start_row: int, end_row: int, date_col: int) -> tuple[set[str], pd.Timestamp | None]:
    existing_dates: set[str] = set()
    max_date: pd.Timestamp | None = None

    for row in range(start_row + 1, end_row + 1):
        raw_date = sheet.cell(row=row, column=date_col).value
        parsed_date = pd.to_datetime(raw_date, errors="coerce")
        if pd.isna(parsed_date):
            continue
        day = parsed_date.normalize()
        day_key = day.strftime("%Y-%m-%d")
        existing_dates.add(day_key)
        if max_date is None or day > max_date:
            max_date = day

    return existing_dates, max_date


def _strip_accents(text: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))


def normalize_description(text: object) -> str:
    value = _strip_accents(str(text).strip().lower())
    value = re.sub(r"^(pago|transferencia enviada|transferencia recibida)\s+", "", value)
    value = re.sub(r"[^a-z0-9\s]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _build_compound_key(date_value, monto_value, tipo_value, desc_value) -> tuple[str, float, str, str]:
    date = pd.to_datetime(date_value, errors="coerce")
    day = date.strftime("%Y-%m-%d") if pd.notna(date) else ""
    if pd.notna(monto_value):
        try:
            monto = round(float(monto_value), 2)
        except (TypeError, ValueError):
            monto = round(parse_number_ar(monto_value), 2)
    else:
        monto = 0.0
    tipo = str(tipo_value).strip()
    desc_norm = normalize_description(desc_value)
    return (day, monto, tipo, desc_norm)


def _extract_existing_compound_keys(
    sheet,
    start_row: int,
    end_row: int,
    date_col: int,
    tipo_col: int,
    descripcion_col: int,
    monto_col: int,
) -> set[tuple[str, float, str, str]]:
    keys: set[tuple[str, float, str, str]] = set()
    for row in range(start_row + 1, end_row + 1):
        fecha = sheet.cell(row=row, column=date_col).value
        tipo = sheet.cell(row=row, column=tipo_col).value
        descripcion = sheet.cell(row=row, column=descripcion_col).value
        monto = sheet.cell(row=row, column=monto_col).value
        key = _build_compound_key(fecha, monto, tipo, descripcion)
        if key[0]:
            keys.add(key)
    return keys


def _resolve_amount_format(sheet, template_row: int, amount_col: int) -> str:
    fmt = sheet.cell(row=template_row, column=amount_col).number_format
    return fmt if fmt else DEFAULT_ARS_FORMAT


def _mes_formula_for_row(row: int, date_col: int) -> str:
    date_col_letter = _index_to_col(date_col)
    return f"=DATE(YEAR({date_col_letter}{row}),MONTH({date_col_letter}{row}),1)"


def _copy_row_style(sheet, src_row: int, dst_row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        src_cell = sheet.cell(row=src_row, column=col)
        dst_cell = sheet.cell(row=dst_row, column=col)
        dst_cell._style = copy(src_cell._style)
        if src_cell.has_style:
            dst_cell.number_format = src_cell.number_format
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.protection = copy(src_cell.protection)


def _normalize_day_key(value: object) -> str:
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return ""
    return parsed.normalize().strftime("%Y-%m-%d")


def _apply_date_filter(movements_df: pd.DataFrame, existing_dates: set[str], max_date: pd.Timestamp | None, mode: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    candidate = movements_df.copy()
    if mode == DATE_FILTER_MODE_AFTER_MAX:
        if max_date is None:
            return candidate, candidate.iloc[0:0].copy()
        date_series = pd.to_datetime(candidate["date"], errors="coerce").dt.normalize()
        keep_mask = date_series > max_date
        return candidate.loc[keep_mask].copy(), candidate.loc[~keep_mask].copy()

    if mode == DATE_FILTER_MODE_SKIP_EXISTING:
        day_keys = candidate["date"].map(_normalize_day_key)
        keep_mask = ~day_keys.isin(existing_dates)
        return candidate.loc[keep_mask].copy(), candidate.loc[~keep_mask].copy()

    raise ValueError(f"Modo de filtro por fecha no soportado: {mode}")


def build_import_plan(
    movements_df: pd.DataFrame,
    finanzas_path: Path,
    date_filter_mode: str = DATE_FILTER_MODE_AFTER_MAX,
) -> ImportPlan:
    if not finanzas_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {finanzas_path}")

    wb = load_workbook(finanzas_path, data_only=False)
    resolved_sheet_name = resolve_sheet_name(wb)
    if resolved_sheet_name is None:
        raise ValueError(f"No existe una hoja compatible en el workbook. Se esperaba {SHEET_NAME}.")

    ws = wb[resolved_sheet_name]
    target = _resolve_target_range(ws)
    start_col, start_row, end_row = target.start_col, target.start_row, target.end_row
    headers = _build_header_map(ws, start_row, start_col, target.end_col)
    date_col = headers["fecha"]
    tipo_col = headers["tipo"]
    descripcion_col = headers["descripcion"]
    monto_col = headers["monto"]
    notes_col = headers.get("notas")

    existing_refs = _extract_existing_refs(ws, start_row, end_row, notes_col)
    existing_keys = _extract_existing_compound_keys(ws, start_row, end_row, date_col, tipo_col, descripcion_col, monto_col)
    existing_dates, max_date = _extract_existing_dates(ws, start_row, end_row, date_col)

    date_candidates, filtered_by_date_df = _apply_date_filter(
        movements_df=movements_df,
        existing_dates=existing_dates,
        max_date=max_date,
        mode=date_filter_mode,
    )

    mp_ref_mask = date_candidates["mp_ref"].astype(str).str.strip().isin(existing_refs)
    duplicate_mp_ref_df = date_candidates.loc[mp_ref_mask].copy()
    candidates = date_candidates.loc[~mp_ref_mask].copy()

    keys = candidates.apply(
        lambda row: _build_compound_key(row["date"], row["monto"], row["tipo"], row["descripcion"]),
        axis=1,
    )
    key_mask = keys.isin(existing_keys)
    duplicate_compound_df = candidates.loc[key_mask].copy()
    to_import_df = candidates.loc[~key_mask].copy()

    wb.close()
    return ImportPlan(
        to_import_df=to_import_df.reset_index(drop=True),
        filtered_by_date_df=filtered_by_date_df.reset_index(drop=True),
        duplicate_mp_ref_df=duplicate_mp_ref_df.reset_index(drop=True),
        duplicate_compound_df=duplicate_compound_df.reset_index(drop=True),
        existing_refs=existing_refs,
        total_input_rows=int(movements_df.shape[0]),
        table_name=target.table_name,
        date_filter_mode=date_filter_mode,
    )


def import_into_finanzas_workbook(
    movements_df: pd.DataFrame,
    finanzas_path: Path,
    date_filter_mode: str = DATE_FILTER_MODE_AFTER_MAX,
    default_category: str = "Otros",
) -> ImportResult:
    if not finanzas_path.exists():
        raise FileNotFoundError(f"No existe el archivo: {finanzas_path}")

    plan = build_import_plan(movements_df, finanzas_path, date_filter_mode=date_filter_mode)
    if plan.to_import_df.empty:
        return ImportResult(
            added_rows=0,
            duplicate_rows_mp_ref=int(plan.duplicate_mp_ref_df.shape[0]),
            duplicate_rows_compound_key=int(plan.duplicate_compound_df.shape[0]),
            skipped_rows_by_date=int(plan.filtered_by_date_df.shape[0]),
            backup_path=None,
            note="No rows to import",
        )

    backup_path = create_backup_path(finanzas_path)
    source_bytes = finanzas_path.read_bytes()
    try:
        backup_path.write_bytes(source_bytes)
    except PermissionError:
        temp_backup = Path(gettempdir()) / backup_path.name
        temp_backup.write_bytes(source_bytes)
        backup_path = temp_backup

    wb = load_workbook(finanzas_path)
    resolved_sheet_name = resolve_sheet_name(wb)
    if resolved_sheet_name is None:
        raise ValueError(f"No existe una hoja compatible en el workbook. Se esperaba {SHEET_NAME}.")

    ws = wb[resolved_sheet_name]
    target = _resolve_target_range(ws)
    start_col, start_row, end_col, end_row = target.start_col, target.start_row, target.end_col, target.end_row
    headers = _build_header_map(ws, start_row, start_col, end_col)
    date_col = headers["fecha"]
    tipo_col = headers["tipo"]
    categoria_col = headers["categoria"]
    subcategoria_col = headers["subcategoria"]
    descripcion_col = headers["descripcion"]
    monto_col = headers["monto"]
    cuenta_col = headers["cuenta"]
    canal_col = headers["canal"]
    mes_col = headers["mes"]
    compartido_col = headers["compartido"]
    notes_col = headers.get("notas")
    existing_refs = set(plan.existing_refs)

    added = 0

    current_end = end_row
    odd_template_row = end_row if end_row <= start_row + 1 else end_row - 1
    even_template_row = end_row
    amount_format = _resolve_amount_format(ws, end_row, monto_col)

    for _, mov in plan.to_import_df.iterrows():
        ref = str(mov["mp_ref"]).strip()
        if not ref or ref in existing_refs:
            continue

        new_row = current_end + 1
        template_row = odd_template_row if (new_row - (end_row + 1)) % 2 == 0 else even_template_row
        _copy_row_style(ws, template_row, new_row, start_col, end_col)
        clean_desc = str(mov["descripcion"]).strip()
        categoria = str(mov["categoria"]).strip() if "categoria" in mov and pd.notna(mov["categoria"]) else default_category
        subcategoria = str(mov["subcategoria"]).strip() if "subcategoria" in mov and pd.notna(mov["subcategoria"]) else ""
        compartido = str(mov["compartido"]).strip() if "compartido" in mov and pd.notna(mov["compartido"]) else "No"
        source_account = (
            str(mov["source_account"]).strip()
            if "source_account" in mov and pd.notna(mov["source_account"])
            else "Mercado Pago"
        )
        source_channel = (
            str(mov["source_channel"]).strip()
            if "source_channel" in mov and pd.notna(mov["source_channel"])
            else "General"
        )

        date_value = pd.to_datetime(mov["date"], errors="coerce")
        date_cell = ws.cell(row=new_row, column=date_col, value=date_value.date() if pd.notna(date_value) else None)
        date_cell.number_format = DEFAULT_DATE_FORMAT
        ws.cell(row=new_row, column=tipo_col, value=mov["tipo"])
        ws.cell(row=new_row, column=categoria_col, value=categoria)
        ws.cell(row=new_row, column=subcategoria_col, value=subcategoria)
        ws.cell(row=new_row, column=descripcion_col, value=clean_desc)
        amount_cell = ws.cell(row=new_row, column=monto_col, value=float(mov["monto"]))
        amount_cell.number_format = amount_format
        ws.cell(row=new_row, column=cuenta_col, value=source_account)
        ws.cell(row=new_row, column=canal_col, value=source_channel)
        mes_cell = ws.cell(row=new_row, column=mes_col, value=_mes_formula_for_row(new_row, date_col))
        mes_cell.number_format = DEFAULT_MES_FORMAT
        source_note = ""
        if "source_note" in mov and pd.notna(mov["source_note"]):
            source_note = str(mov["source_note"]).strip()
        note_value = source_note if source_note else (f"mp_ref={ref}" if ref else None)
        if notes_col is not None:
            ws.cell(row=new_row, column=notes_col, value=note_value)
        ws.cell(row=new_row, column=compartido_col, value=compartido)

        existing_refs.add(ref)
        current_end = new_row
        added += 1

    # Extiende el rango de la tabla hasta la nueva ultima fila (si existe tabla).
    table = _get_table(ws)
    if table is not None and current_end != end_row:
        start_col_letter = _index_to_col(start_col)
        end_col_letter = _index_to_col(end_col)
        table.ref = f"{start_col_letter}{start_row}:{end_col_letter}{current_end}"
        if table.tableStyleInfo is not None:
            table.tableStyleInfo.showRowStripes = True

    try:
        wb.save(finanzas_path)
    except PermissionError as exc:
        wb.close()
        raise PermissionError(
            "Cerrá FINANZAS.xlsx en Excel y/o marcá el archivo como Disponible sin conexión en Google Drive."
        ) from exc
    wb.close()

    return ImportResult(
        added_rows=added,
        duplicate_rows_mp_ref=int(plan.duplicate_mp_ref_df.shape[0]),
        duplicate_rows_compound_key=int(plan.duplicate_compound_df.shape[0]),
        skipped_rows_by_date=int(plan.filtered_by_date_df.shape[0]),
        backup_path=backup_path,
    )
