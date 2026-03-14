from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from .workbook_writer import SHEET_NAME, _resolve_target_range, resolve_sheet_name
from .utils import parse_number_ar


@dataclass
class MonthlyKpis:
    month_label: str
    ingresos: float
    gastos: float
    balance: float
    ahorro_pct: float
    gasto_fijo_aprox: float
    gasto_variable_aprox: float
    top_categories: pd.DataFrame
    delta_balance_prev: float
    delta_balance_avg3: float
    delta_gasto_prev: float
    delta_gasto_avg3: float


@dataclass
class AlertItem:
    level: str  # critical | warning | info
    title: str
    detail: str
    action: str


def _safe_numeric(series: pd.Series) -> pd.Series:
    return series.map(parse_number_ar).astype(float)


def load_finanzas_history(finanzas_path: Path) -> pd.DataFrame:
    if not finanzas_path.exists():
        return pd.DataFrame(columns=["date", "tipo", "categoria", "descripcion", "monto"])

    wb = load_workbook(finanzas_path, data_only=True)
    resolved_sheet_name = resolve_sheet_name(wb)
    if resolved_sheet_name is None:
        wb.close()
        return pd.DataFrame(columns=["date", "tipo", "categoria", "descripcion", "monto"])

    ws = wb[resolved_sheet_name]
    target = _resolve_target_range(ws)
    start_col, start_row, end_row = target.start_col, target.start_row, target.end_row

    rows: list[dict[str, object]] = []
    for row in range(start_row + 1, end_row + 1):
        date_value = ws.cell(row=row, column=start_col + 0).value
        tipo = ws.cell(row=row, column=start_col + 1).value
        categoria = ws.cell(row=row, column=start_col + 2).value
        descripcion = ws.cell(row=row, column=start_col + 4).value
        monto = ws.cell(row=row, column=start_col + 5).value
        if date_value is None or monto is None:
            continue
        rows.append(
            {
                "date": date_value,
                "tipo": str(tipo).strip() if tipo is not None else "",
                "categoria": str(categoria).strip() if categoria is not None else "Otros",
                "descripcion": str(descripcion).strip() if descripcion is not None else "",
                "monto": monto,
            }
        )

    wb.close()

    if not rows:
        return pd.DataFrame(columns=["date", "tipo", "categoria", "descripcion", "monto"])

    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"]).copy()
    df["monto"] = _safe_numeric(df["monto"])
    df["tipo"] = df["tipo"].replace("", "Gasto")
    df["signed_monto"] = df.apply(lambda r: -abs(r["monto"]) if str(r["tipo"]).lower() == "gasto" else abs(r["monto"]), axis=1)
    return df.sort_values("date").reset_index(drop=True)


def merge_history_with_pending(history_df: pd.DataFrame, pending_df: pd.DataFrame | None) -> pd.DataFrame:
    if pending_df is None or pending_df.empty:
        return history_df.copy()

    pending = pending_df.copy()
    required_cols = {"date", "tipo", "categoria", "descripcion", "monto"}
    if not required_cols.issubset(pending.columns):
        return history_df.copy()

    pending["date"] = pd.to_datetime(pending["date"], errors="coerce")
    pending = pending.dropna(subset=["date"]).copy()
    pending["monto"] = _safe_numeric(pending["monto"])
    pending["signed_monto"] = pending.apply(
        lambda r: -abs(r["monto"]) if str(r["tipo"]).lower() == "gasto" else abs(r["monto"]),
        axis=1,
    )
    pending = pending[["date", "tipo", "categoria", "descripcion", "monto", "signed_monto"]]
    pending["source"] = "pending_import"

    combined = history_df.copy()
    if "source" not in combined.columns:
        combined["source"] = "finanzas"
    return pd.concat([combined, pending], ignore_index=True).sort_values("date").reset_index(drop=True)


def _month_window(df: pd.DataFrame, year: int, month: int) -> pd.DataFrame:
    return df[(df["date"].dt.year == year) & (df["date"].dt.month == month)].copy()


def compute_monthly_kpis(df: pd.DataFrame, ref_date: pd.Timestamp | None = None) -> MonthlyKpis | None:
    if df.empty:
        return None

    ref = ref_date if ref_date is not None else df["date"].max()
    current = _month_window(df, ref.year, ref.month)
    if current.empty:
        return None

    prev_ref = (ref.replace(day=1) - pd.Timedelta(days=1))
    prev = _month_window(df, prev_ref.year, prev_ref.month)

    month_starts = sorted(df["date"].dt.to_period("M").dropna().unique())
    recent_periods = [m for m in month_starts if m != ref.to_period("M")]
    recent_periods = recent_periods[-3:] if len(recent_periods) >= 3 else recent_periods
    recent_df = df[df["date"].dt.to_period("M").isin(recent_periods)].copy()

    ingresos = float(current.loc[current["signed_monto"] > 0, "signed_monto"].sum())
    gastos = float(abs(current.loc[current["signed_monto"] < 0, "signed_monto"].sum()))
    balance = ingresos - gastos
    ahorro_pct = (balance / ingresos * 100.0) if ingresos > 0 else 0.0

    by_desc = (
        current.loc[current["signed_monto"] < 0]
        .groupby("descripcion", as_index=False)["signed_monto"]
        .sum()
        .rename(columns={"signed_monto": "gasto"})
    )
    gasto_fijo = float(abs(by_desc.loc[by_desc["gasto"] < -1.0, "gasto"].head(5).sum()))
    gasto_variable = max(gastos - gasto_fijo, 0.0)

    top_categories = (
        current.loc[current["signed_monto"] < 0]
        .groupby("categoria", as_index=False)["signed_monto"]
        .sum()
        .assign(gasto=lambda d: d["signed_monto"].abs())
        .sort_values("gasto", ascending=False)
        .head(5)[["categoria", "gasto"]]
    )

    prev_balance = float(prev["signed_monto"].sum()) if not prev.empty else 0.0
    prev_gasto = float(abs(prev.loc[prev["signed_monto"] < 0, "signed_monto"].sum())) if not prev.empty else 0.0
    recent_balance_avg = float(
        recent_df.groupby(recent_df["date"].dt.to_period("M"))["signed_monto"].sum().mean()
    ) if not recent_df.empty else 0.0
    recent_gasto_avg = float(
        recent_df.loc[recent_df["signed_monto"] < 0]
        .groupby(recent_df.loc[recent_df["signed_monto"] < 0, "date"].dt.to_period("M"))["signed_monto"]
        .sum()
        .abs()
        .mean()
    ) if not recent_df.empty else 0.0

    return MonthlyKpis(
        month_label=ref.strftime("%B %Y"),
        ingresos=ingresos,
        gastos=gastos,
        balance=balance,
        ahorro_pct=ahorro_pct,
        gasto_fijo_aprox=gasto_fijo,
        gasto_variable_aprox=gasto_variable,
        top_categories=top_categories.reset_index(drop=True),
        delta_balance_prev=balance - prev_balance,
        delta_balance_avg3=balance - recent_balance_avg,
        delta_gasto_prev=gastos - prev_gasto,
        delta_gasto_avg3=gastos - recent_gasto_avg,
    )


def compute_month_projection(df: pd.DataFrame, ref_date: pd.Timestamp | None = None) -> dict[str, float] | None:
    if df.empty:
        return None
    ref = ref_date if ref_date is not None else df["date"].max()
    current = _month_window(df, ref.year, ref.month)
    if current.empty:
        return None

    gastos_actuales = float(abs(current.loc[current["signed_monto"] < 0, "signed_monto"].sum()))
    ingresos_actuales = float(current.loc[current["signed_monto"] > 0, "signed_monto"].sum())

    observed_day = int(current["date"].dt.day.max()) if not current.empty else int(ref.day)
    days_elapsed = max(observed_day, 1)
    days_in_month = int((ref + pd.offsets.MonthEnd(0)).day)
    factor = days_in_month / days_elapsed

    gasto_proyectado = gastos_actuales * factor
    ingreso_proyectado = ingresos_actuales * factor
    balance_proyectado = ingreso_proyectado - gasto_proyectado
    return {
        "days_elapsed": float(days_elapsed),
        "days_in_month": float(days_in_month),
        "gasto_proyectado": float(gasto_proyectado),
        "ingreso_proyectado": float(ingreso_proyectado),
        "balance_proyectado": float(balance_proyectado),
    }


def build_alerts(
    df: pd.DataFrame,
    projection: dict[str, float] | None,
    ref_date: pd.Timestamp | None = None,
    ahorro_meta_pct: float = 10.0,
) -> list[AlertItem]:
    alerts: list[AlertItem] = []
    if df.empty:
        return alerts

    ref = ref_date if ref_date is not None else df["date"].max()
    current = _month_window(df, ref.year, ref.month)
    if current.empty:
        return alerts

    ingresos_mes = float(current.loc[current["signed_monto"] > 0, "signed_monto"].sum())
    gastos_mes = float(abs(current.loc[current["signed_monto"] < 0, "signed_monto"].sum()))
    ahorro_pct = ((ingresos_mes - gastos_mes) / ingresos_mes * 100.0) if ingresos_mes > 0 else 0.0

    if projection is not None and projection["balance_proyectado"] < 0:
        alerts.append(
            AlertItem(
                level="critical",
                title="Riesgo de cierre negativo",
                detail=f"El balance proyectado cierra en ARS {projection['balance_proyectado']:,.0f}.",
                action="Reducir gasto variable esta semana y postergar compras no esenciales.",
            )
        )

    if ahorro_pct < ahorro_meta_pct:
        alerts.append(
            AlertItem(
                level="warning",
                title="Ahorro por debajo de objetivo",
                detail=f"Ahorro actual {ahorro_pct:.1f}% vs meta {ahorro_meta_pct:.1f}%.",
                action="Definir tope diario de gasto hasta fin de mes para recuperar margen.",
            )
        )

    by_cat_current = (
        current.loc[current["signed_monto"] < 0]
        .groupby("categoria", as_index=False)["signed_monto"]
        .sum()
        .assign(gasto=lambda d: d["signed_monto"].abs())
    )
    prev_ref = (ref.replace(day=1) - pd.Timedelta(days=1))
    prev = _month_window(df, prev_ref.year, prev_ref.month)
    by_cat_prev = (
        prev.loc[prev["signed_monto"] < 0]
        .groupby("categoria", as_index=False)["signed_monto"]
        .sum()
        .assign(gasto_prev=lambda d: d["signed_monto"].abs())[["categoria", "gasto_prev"]]
    )
    merged = by_cat_current.merge(by_cat_prev, on="categoria", how="left").fillna(0.0)
    merged["ratio"] = merged.apply(
        lambda r: (r["gasto"] / r["gasto_prev"]) if r["gasto_prev"] > 0 else 0.0,
        axis=1,
    )
    spikes = merged[(merged["gasto"] >= 1000) & (merged["ratio"] >= 1.30)].sort_values("ratio", ascending=False)
    for _, row in spikes.head(3).iterrows():
        alerts.append(
            AlertItem(
                level="warning",
                title=f"Suba fuerte en {row['categoria']}",
                detail=f"+{((row['ratio'] - 1) * 100):.0f}% vs mes anterior (ARS {row['gasto']:,.0f}).",
                action=f"Revisar consumos de {row['categoria']} y fijar limite semanal.",
            )
        )

    if current.shape[0] > 0:
        day_spend = current.loc[current["signed_monto"] < 0].copy()
        if not day_spend.empty:
            daily = day_spend.groupby(day_spend["date"].dt.day)["signed_monto"].sum().abs()
            threshold = daily.mean() + (2 * daily.std(ddof=0) if daily.shape[0] > 1 else 0.0)
            peaks = daily[daily > threshold]
            if not peaks.empty:
                day = int(peaks.sort_values(ascending=False).index[0])
                amount = float(peaks.max())
                alerts.append(
                    AlertItem(
                        level="info",
                        title=f"Pico de gasto en dia {day}",
                        detail=f"Se detecto un gasto diario atipico de ARS {amount:,.0f}.",
                        action="Validar si fue gasto excepcional o un nuevo patron recurrente.",
                    )
                )

    if not alerts:
        alerts.append(
            AlertItem(
                level="info",
                title="Sin alertas criticas",
                detail="No se detectaron desvios relevantes para este periodo.",
                action="Mantener el ritmo actual y controlar categorias variables.",
            )
        )
    return alerts


def build_executive_summary(alerts: list[AlertItem], kpis: MonthlyKpis | None, projection: dict[str, float] | None) -> str:
    if kpis is None:
        return "Sin datos suficientes para resumen ejecutivo."
    critical = sum(1 for a in alerts if a.level == "critical")
    warning = sum(1 for a in alerts if a.level == "warning")
    projected = projection["balance_proyectado"] if projection is not None else kpis.balance
    return (
        f"Balance mes: ARS {kpis.balance:,.0f}. "
        f"Cierre proyectado: ARS {projected:,.0f}. "
        f"Alertas: {critical} criticas, {warning} advertencias."
    )
